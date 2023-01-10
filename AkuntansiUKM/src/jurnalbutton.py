import datetime 
import tkinter as tk
from tkinter.ttk import *
import tkinter.ttk as ttk
from tkinter import *
from tkinter import ttk, messagebox
from informasialert import Pesan
from openpyxl import Workbook 
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import json, configsql, updateframedata, os
from datetime import date

class Jurnal():
    def __init__(self, roots):
        self.roots = roots
        self.tambahjurnal()

    def tambahjurnal(self):
        today = date.today()
        self.d2 = today.strftime("%B %d, %Y")
        print("Today is ", self.d2)

        transaksi = Toplevel(self.roots)
        transaksi.title("Transaksi Baru")
        transaksi.geometry("430x227")

        # pembeda
        # pemasukan
        # jika tab pemasukan, label menjadi 'diterima dari' dan 'simpan ke' 
        labelditerimadari = "Diterima Dari"
        combopemasukanditerimadari = [
                            'Pendapatan', 
                            'Penjualan Barang', 
                            'Ikhtisar Laba/Rugi',
                            'Potongan Penjualan',
                            'Retur Penjualan',
                            'Pendapatan Diluar Usaha',
                            'Pendapatan Bunga Bank',
                            'Pendapatan Hasil Panen'
                            ]
        combopemasukansimpanke = [
                            'Kas',
                            'Bank',
                            'Perlengkapan',
                            'Persediaan Bahan Baku',
                            'Persediaan Barang Dagang',
                            'Sewa Dibayar Dimuka'
                            ]

        # pengeluaran
        # jika tab pengeluaran, label menjadi 'diambil dari' dan 'untuk'
        labeldiambildari = 'Diambil Dari' 
        combopengeluarandiambildari = [
                            'Kas', 
                            'Bank', 
                            'Perlengkapan', 
                            'Persediaan Bahan Baku', 
                            'Persediaan Barang Dagang', 
                            'Sewa Dibayar Dimuka', 
                            'Tanah', 
                            'Bangunan', 
                            'Kendaraan', 
                            'Peralatan'
                            ]
        labeluntuk = 'Untuk'
        combopengeluaranuntuk = [
                            'Biaya Air', 
                            'Biaya Gaji Karyawan', 
                            'Biaya Listrik', 
                            'Biaya Makan dan Minum', 
                            'Biaya Perlengkapan', 
                            'Biaya Sewa Tempat Usaha', 
                            'Biaya Telpon', 
                            'Biaya Umum Lain-lain', 
                            'Beban Penyusutan Peralatan', 
                            'Beban Piutang Tak Tertagih', 
                            'Biaya Pengiriman', 
                            'Biaya Penjualan Lain-lain', 
                            'Biaya Administrasi Bank', 
                            'Harga Pokok Penjualan', 
                            'Potongan Pembelian', 
                            'Retur Pembelian', 
                            'Tanah', 
                            'Bangunan', 
                            'Kendaraan', 
                            'Peralatan', 
                            'Kas', 
                            'Bank',
                            'Perlengkapan',
                            'Persediaan Bahan Baku',
                            'Persediaan Barang Dagang',
                            'Piutang Usaha',
                            'Sewa Dibayar Dimuka'
                            ]

        # jika tab utang, label menjadi 'Utang dari' dan 'Simpan ke' 
        labelutangdari = 'Utang Dari'
        comboutangutangdari = [
                            'Utang Usaha',
                            'Utang Bank'
                            ]
        labelsimpanke = 'Simpan ke'
        comboutangsimpanke = [
                            'Kas',
                            'Bank',
                            'Perlengkapan',
                            'Persediaan Bahan Baku',
                            'Persediaan Barang Dagang',
                            'Sewa Dibayar Dimuka',
                            'Tanah',
                            'Bangunan',
                            'Kendaraan',
                            'Peralatan',
                            'Biaya Pengiriman',
                            'Biaya Penjualan Lain-lain',
                            'Biaya Air',
                            'Biaya Depresiasi Peralatan',
                            'Biaya Gaji Karyawan',
                            'Biaya Listrik',
                            'Biaya Makan dan Minum',
                            'Biaya Perlengkapan',
                            'Biaya Sewa Tempat Usaha',
                            'Biaya Telpon',
                            'Biaya Umum Lain-lain',
                            'Beban Penyusutan Bangunan',
                            'Beban Penyusutan Kendaraan',
                            'Beban Penyusutan Peralatan',
                            'Biaya Administrasi Bank'
                            ]

        # jika tab utang, label menjadi 'Utang dari' dan 'Simpan ke' 
        comboutambahmodal = [
                            'Modal Pemilik',
                            'Prive'
                            ]


        # Penyesuaian
        # jika tab pengeluaran, label menjadi 'diambil dari' dan 'untuk'
        combopenyesuaianke = [
                                'Kas',
                                'Bank',
                                'Perlengkapan',
                                'Persediaan Bahan Baku',
                                'Persediaan Barang Dagang',
                                'Piutang Usaha',
                                'Sewa Dibayar Dimuka',
                                'Tanah',
                                'Bangunan',
                                'Kendaraan',
                                'Peralatan',
                                'Akumulasi Penyusutan Kendaraan',
                                'Akumulasi Penyusutan Peralatan',
                                'Akumulasi Penyusutan Bangunan',
                                'Utang Usaha',
                                'Utang Bank',
                                'Pendapatan',
                                'Penjualan Barang',
                                'Ikhtisar Laba/Rugi',
                                'Potongan Penjualan',
                                'Retur Penjualan',
                                'Harga Pokok Penjualan',
                                'Potongan Pembelian',
                                'Biaya Pengiriman',
                                'Biaya Penjualan Lain-lain',
                                'Biaya Air',
                                'Biaya Gaji Karyawan',
                                'Biaya Listrik',
                                'Biaya Makan dan Minum',
                                'Biaya Perlegkapan',
                                'Biaya Sewa Tempat Usaha',
                                'Biaya Telpon',
                                'Biaya Umum Lain-lain',
                                'Beban Penyusutan Bangunan',
                                'Beban Penyusutan Kendaraan',
                                'Beban Penyusutan Peralatan',
                                'Beban Piutang Tak Tertagih',
                                'Pendapatan Bunga Bank',
                                'Pendapatan Hasil Panen',
                                'Biaya Administrasi Bank'
                            ]


        # transaksi.resizable(False, False)

        #judul
        Label(transaksi, text="Tambah Data Transaksi", font="courier,bold, 12").grid(row=0, column=0, columnspan=1, pady=5, padx=10)

        #pemasukan
        self.varpemasukan = StringVar()
        Label(transaksi, text=self.d2, font="courier, 10").grid(row=1, column=0,  pady=5, padx=10, sticky=W)
        self.datalist = ['Pemasukan','Pengeluaran','Utang','Bayar Utang','Piutang','Dibayar Piutang','Tambah Modal','Tarik Modal','Pengalihan Asset','Penyesuaian']
        combotransaksi = ttk.Combobox(transaksi, state='readonly', textvariable= self.varpemasukan, values=self.datalist)
        combotransaksi.grid(row = 1, column=1, columnspan=1, sticky=W)
        combotransaksi.current(0)

        def penentuan():
            if self.varpemasukan.get() == 'Pemasukan':
                self.labela.config(text='Diterima Dari')
                self.combosatu['values'] = combopemasukanditerimadari
                self.combosatu.current(0)
                self.labelb.config(text='Simpan Ke')
                self.combodua['values'] = combopemasukansimpanke
                self.combodua.current(0)
            elif self.varpemasukan.get() == 'Pengeluaran':
                self.labela.config(text='Diambil Dari')
                self.combosatu['values'] = combopengeluarandiambildari
                self.combosatu.current(0)
                self.labelb.config(text='Untuk')
                self.combodua['values'] = combopengeluaranuntuk
                self.combodua.current(0)
            elif self.varpemasukan.get() == 'Utang':
                self.labela.config(text='Utang Dari')
                self.combosatu['values'] = comboutangutangdari
                self.combosatu.current(0)
                self.labelb.config(text='Simpan Ke')
                self.combodua['values'] = comboutangsimpanke
                self.combodua.current(0)
            
            elif self.varpemasukan.get() == 'Bayar Utang':
                toplevelbayarhutang()

            elif self.varpemasukan.get() == 'Tarik Modal':
                self.labela.config(text='Diambil Dari')
                self.combosatu['values'] = combopengeluarandiambildari
                self.combosatu.current(0)
                self.labelb.config(text='Modal')
                self.combodua['values'] = comboutambahmodal
                self.combodua.current(0)
            elif self.varpemasukan.get() == 'Tambah Modal':
                self.labela.config(text='Modal')
                self.combosatu['values'] = comboutambahmodal
                self.combosatu.current(0)
                self.labelb.config(text='Simpan Ke')
                self.combodua['values'] = combopengeluarandiambildari
                self.combodua.current(0)
            elif self.varpemasukan.get() == 'Pengalihan Asset':
                self.labela.config(text='Dari')
                self.combosatu['values'] = combopengeluarandiambildari
                self.combosatu.current(0)
                self.labelb.config(text='Ke')
                self.combodua['values'] = combopengeluarandiambildari
                self.combodua.current(0)
            elif self.varpemasukan.get() == 'Penyesuaian':
                self.labela.config(text='Dari')
                self.combosatu['values'] = combopenyesuaianke
                self.combosatu.current(0)
                self.labelb.config(text='Ke')
                self.combodua['values'] = combopenyesuaianke
                self.combodua.current(0)

        tombolpilih = ttk.Button(transaksi, text='@', width=2, command=lambda: penentuan())
        tombolpilih.grid(row=1, column=2)
        self.tmp = ()
        #diterima dari

        #Keterangan
        Label(transaksi, text="Deskripsi", font="courier, 10").grid(row=4, column=0, pady=5, padx=10, sticky=W)
        self.keterangan = ttk.Entry(transaksi, width=22)
        self.getketerangan = self.keterangan.get()
        self.keterangan.grid(row=4, column=1)

        #nominal
        Label(transaksi, text="Nominal", font="courier, 10").grid(row=5, column=0, pady=5, padx=10, sticky=W)
        self.nominal = ttk.Entry(transaksi, width=22)
        self.getnominal = self.nominal.get()
        self.nominal.grid(row=5, column=1)

        def debet():
            self.varditerimadari = StringVar()
            self.labela = Label(transaksi, text="Diterima dari", font="courier, 10")
            self.labela.grid(row=2, column=0,  pady=5, padx=10, sticky=W)
            self.combosatu = ttk.Combobox(transaksi, state='readonly', textvariable=self.varditerimadari ,values=['Pendapatan','Penjualan Barang','Ikhtisar Laba/Rugi','Potongan Penjualan','Retur Penjualan','Pendapatan Diluar Usaha','Pendapatan Bunga Bank','Pendapatan Hasil Panen'])
            self.combosatu.grid(row = 2, column=1, sticky=W, columnspan=1)
            self.combosatu.current(0)
        debet()

        #Simpan ke
        def kredit():
            self.varsimpanke = StringVar()
            self.labelb = Label(transaksi, text="Simpan ke", font="courier, 10")
            self.labelb.grid(row=3, column=0,  pady=5, padx=10, sticky=W)
            self.combodua = ttk.Combobox(transaksi, state='readonly', textvariable=self.varsimpanke ,values=['Aktiva Lancar','Kas','Bank','Perlengkapan','Persediaan Bahan Baku','Persediaan Barang Dagang','Sewa Dibayar Dimuka'])
            self.combodua.grid(row = 3, column=1, sticky=W, columnspan=1)
            self.combodua.current(0)
        kredit()

        print(self.keterangan.get())
        #tombol simpan
        tombolsimpan = ttk.Button(transaksi, text="Simpan", width=27, command=lambda: [self.savedatabase(), self.savedjson(), Pesan(transaksi).sukses(), self.refreshgui()]).grid(row=6, columnspan=3, column=0, pady=5)       


        def toplevelbayarhutang(*args):
            print('Option 3 is selected')
            # event.widget.get()
            today = datetime.date.today()
            year = today.year
            utangg = Toplevel(transaksi)
            utangg.title("Utang")
            utangg.geometry("430x250")

            Label(utangg, text="Bayar Utang", font=('Calibri 15')).grid(row=0, column=0)
            combotahunutang = ttk.Combobox(utangg, state='readonly', values=year)
            combotahunutang.grid(row=1, column=0, padx=15)
            combotahunutang.current(0)
            entrycarihutang = ttk.Entry(utangg)
            entrycarihutang.grid(row=1, column=1)
            tombolcari = ttk.Button(utangg, text='Cari', width=3)
            tombolcari.grid(row=1, column=3, padx=8)

            status = Label(utangg, text="Tidak Ada Data", font=('Calibri 13')).place(relx=0.3, rely=0.5)


    def refreshgui(self):
        try:
            self.roots.destroy()
            os.popen("/home/python/Tugas/AkuntansiUKM/bin/python /home/python/Tugas/AkuntansiUKM/src/app.py")
        except:
            print("Error geys")

  
    def set_border(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")    
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def savedjson(self):
        # Menulis XLSX File
        today = datetime.datetime.now().strftime('%B')

        # book = Workbook()
        # sheet = book.active

        # redFill = PatternFill(start_color='FFFF0000',
        #         end_color='FFFF0000',
        #         fill_type='solid')

        # # self.set_border(sheet, '')
        
        # sheet['A1'].fill = redFill
        # sheet['A1'].font =  Font(size=14, underline='single', color='FFBB00', bold=True)
        # sheet['A1'] = 'Tanggal'
        # sheet['A2'] = today
        # sheet['A4'] = self.d2
        # sheet['A5'] = self.d2

        # sheet['B1'].fill = redFill
        # sheet['B1'] = 'Keterangan'
        # sheet['B1'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        # sheet['B3'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        # sheet['B3'] = self.getketerangan
        # sheet['B4'] = self.varsimpanke.get()
        # sheet['B5'] = self.varditerimadari.get()

        # sheet['C1'].fill = redFill
        # sheet['C1'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        # sheet['C1'] = 'Debet'
        # sheet['C4'] = self.getnominal
        # sheet['C5'] = self.getnominal

        # sheet['D1'].fill = redFill
        # sheet['D1'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        # sheet['D1'] = 'Kredit'
        # sheet['D4'] = self.getnominal
        # sheet['D5'] = self.getnominal

        # book.save("/home/python/Tugas/AkuntansiUKM/src/documents/excel/sample.xlsx")
        
        #  Menulis Json File
        datajson = {}
        datajson[today] = [{}]
        datajson[today][0]['tanggal'] = self.d2
        datajson[today][0]['keterangan'] = self.varpemasukan.get()
        datajson[today][0]['debit'] = self.varditerimadari.get()
        datajson[today][0]['kredit'] = self.varsimpanke.get()
        datajson[today][0]['deskripsi'] = self.keterangan.get()
        datajson[today][0]['nominal'] = self.nominal.get()

        with open(f'/home/python/Tugas/AkuntansiUKM/src/documents/jsonfolder/{self.keterangan.get()}.json', 'w') as f:
            write = json.dump(datajson, f) 


    def savedxlsx(self):
        from datetime import date
        today = date.today()
        book = Workbook()
        sheet = book.active

        redFill = PatternFill(start_color='FFFF0000',
                end_color='FFFF0000',
                fill_type='solid')

        # self.set_border(sheet, '')
        
        sheet['A1'].fill = redFill
        sheet['A1'].font =  Font(size=14, underline='single', color='FFBB00', bold=True)
        sheet['A1'] = 'Tanggal'
        sheet['A2'] = today
        sheet['A4'] = self.d2
        sheet['A5'] = self.d2

        sheet['B1'].fill = redFill
        sheet['B1'] = 'Keterangan'
        sheet['B1'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        sheet['B3'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        sheet['B3'] = self.getketerangan
        sheet['B4'] = self.varsimpanke.get()
        sheet['B5'] = self.varditerimadari.get()

        sheet['C1'].fill = redFill
        sheet['C1'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        sheet['C1'] = 'Debet'
        sheet['C4'] = self.getnominal
        sheet['C5'] = self.getnominal

        sheet['D1'].fill = redFill
        sheet['D1'].font = Font(size=14, underline='single', color='FFBB00', bold=True)
        sheet['D1'] = 'Kredit'
        sheet['D4'] = self.getnominal
        sheet['D5'] = self.getnominal

        book.save("/home/python/Tugas/AkuntansiUKM/src/documents/excel/sample.xlsx")

        # Store ke Database
    def savedatabase(self):
        today = date.today()
        self.d2 = today.strftime("%B %d, %Y")
        self.tmp = (today, self.keterangan.get(), self.nominal.get(), self.nominal.get())
        configsql.MySQL('localhost', 'python', 'KaliLinux', 'ukm').insert_data('laporanjurnal', self.tmp)  


